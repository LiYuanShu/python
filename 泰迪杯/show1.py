import csv
import datetime
from geopy.distance import geodesic
from geopy.distance import geodesic
import xlwings as xw
import xlwt
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
def route(fileDir):
    ws.title = u'AA00001'
    temp_count_km = 0
    time = 0
    lists = []
    dictt = {}
    llll = [['lng','lat','location_time','Driving mileage','Average speed','Rapid acceleration','count_km']]
    routeDate = []
    v_list = [0]
    time_list = [1]
    with open(fileDir, "r") as csvfile:
        reader = csv.reader(csvfile)
        for j, line in enumerate(reader):
            list4 = []
            if line[5] == '1' and int(line[11]) > 0:
                list4.append(line[3:5])
                list4.append(line[10])
                list4.append(line[11])
            if len(list4) != 0 and len(lists) == 0:
                lists.append(list4)
                routeDate.append(lists[-1][0:2])
                temp_list=[line[3],line[4],line[10],0,0,0,0]
                llll.append(temp_list)
            if len(lists) > 0 and len(list4) != 0:
                if lists[-1][0] != list4[0]:
                    a = lists.append(list4)
                    d1 = datetime.datetime.strptime(lists[-2][1], '%Y-%m-%d %H:%M:%S')
                    d2 = datetime.datetime.strptime(lists[-1][1], '%Y-%m-%d %H:%M:%S')
                    d = d2 - d1
                    h = int(format(d.seconds)) / 3600
                    # print(format(d.seconds))
                    list1 = lists[-2][0]
                    list2 = lists[-1][0]
                    list1 = list(map(float, list1))
                    list2 = list(map(float, list2))
                    list1[0], list1[1] = list1[1], list1[0]
                    list2[0], list2[1] = list2[1], list2[0]
                    tup1 = tuple(list1)
                    tup2 = tuple(list2)
                    # 两个经纬度之间的距离
                    km = geodesic(tup1, tup2).km
                    v = km / h
                    v_list.append(v)
                    a = (v_list[-1] - v_list[-2]) / time_list[-1]
                    temp_count_km = km + temp_count_km
                    time = time + h
                    # routeDate.append(lists[-1][0])
                    distance_v_a = [temp_count_km, v, a]
                    temp_list = [lists[-1][0][0], lists[-1][0][1], lists[-1][1],km, v, a, temp_count_km]
                    llll.append(temp_list)
                    # dictt[str(lists[-1][0:2])] = str(distance_v_a)
    wb.save('test.xlsx')
    return llll


if __name__ == '__main__':
    ff = ['AA00001', 'AB00006', 'AD00003', 'AD00013', 'AD00053', 'AD00083', 'AD00419', 'AF00098', 'AF00131', 'AF00373']
    # route_lists = []
    # for i in ff:
    #     fileDir = 'F:\\pycharm\\shuju\\' + i + '.csv'
    #     route_list = route(fileDir)
    #     dict = {}
    #     dict['name'] = fileDir
    #     dict['path'] = route_list
    #     route_lists.append(dict)
    # with open('F:\\pycharm\\jiqi\\' + 'route.txt', 'a+') as fw:
    #     fw.write(str(route_lists))
    # fw.close()
    route_list = route('F:\\pycharm\\shuju\\AA00001.csv')
    # f = xlwt.Workbook()
    # sheet1 = f.add_sheet(u'sheet2', cell_overwrite_ok=True)
    # print(route_list[-1])
    # for i in range(len(route_list)):
    #     sheet1.write(i, 0, str(route_list[i][0]))
    #     sheet1.write(i, 1, str(route_list[i][1]))
    #     sheet1.write(i, 2, str(route_list[i][2]))
    #     sheet1.write(i, 3, str(route_list[i][3]))
    #     sheet1.write(i, 4, str(route_list[i][4]))
    #     sheet1.write(i, 5, str(route_list[i][5]))
    #     sheet1.write(i, 6, str(route_list[i][6]))
    #
    # f.save(r'C:\Users\asus\Desktop\11.xls')
    wb = Workbook()
    ws = wb.active
    ws.title = u'AA00001'
    i = 1
    r = 1
    for line in route_list:
        for col in range(1, len(line) + 1):
            ColNum = r
            ws.cell(row=r, column=col).value = line[col - 1]
        i += 1
        r += 1
    # 工作簿保存到磁盘
    wb.save('test.xlsx')

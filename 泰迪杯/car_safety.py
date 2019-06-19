import csv
import datetime
from geopy.distance import geodesic
from geopy.distance import geodesic
from interval import Interval
import matplotlib.pyplot as plt
from openpyxl import Workbook
def route(fileDir):
    temp_count_km=0
    time=0
    lists = []
    dictt={}
    routeDate=[]
    fatigue_list=[]
    v_list=[0]
    vv_list=[]
    a_list=[0]
    time_list = [1]
    changLane_list=[]
    flameout_list=[]
    ldleSpeed_list=[]
    ldleSpeed_time_list=[]
    FatigueDriving_list=[]
    FatigueDriving_time_list=[]
    with open(fileDir, "r") as csvfile:
        reader = csv.reader(csvfile)
        for j, line in enumerate(reader):
            list4 = []
            if line[5] == '1' and int(line[11]) != 0 and len(FatigueDriving_list)%2 == 0:
                temp_list = [line[3], line[4],line[10]]
                FatigueDriving_list.append(temp_list)
            if  line[5] == '1' and int(line[11]) == 0 and len(FatigueDriving_list)%2 != 0 :
                temp_list = [line[3], line[4],line[10]]
                FatigueDriving_list.append(temp_list)
                d1 = datetime.datetime.strptime(FatigueDriving_list[-2][2], '%Y-%m-%d %H:%M:%S')
                d2 = datetime.datetime.strptime(FatigueDriving_list[-1][2], '%Y-%m-%d %H:%M:%S')
                d = d2 - d1
                if d.seconds >180:
                    ldleSpeed_time_list.append(d.seconds/60)
            if line[5] == '1' and int(line[11]) == 0 and len(ldleSpeed_list)%2 == 0:
                temp_list = [line[3], line[4],line[10]]
                ldleSpeed_list.append(temp_list)
            if line[5] == '1' and int(line[11]) != 0 and len(ldleSpeed_list)%2 != 0:
                temp_list = [line[3], line[4],line[10]]
                ldleSpeed_list.append(temp_list)
                d1 = datetime.datetime.strptime(ldleSpeed_list[-2][2], '%Y-%m-%d %H:%M:%S')
                d2 = datetime.datetime.strptime(ldleSpeed_list[-1][2], '%Y-%m-%d %H:%M:%S')
                d = d2 - d1
                if d.seconds/3600 >4:
                    FatigueDriving_time_list.append(d.seconds/3600)
            if line[5] == '1' and int(line[11]) > 0:
                list4.append(line[3:5])
                list4.append(line[10])
                list4.append(line[11])
                list4.append(line[2])
            if line[5] == '0' and int(line[11]) > 0:
                temp_list=[line[5],line[11]]
                flameout_list.append(temp_list)
            if len(list4) != 0 and len(lists) == 0:
                lists.append(list4)
                routeDate.append(lists[-1][0:2])
            if len(lists) > 0 and len(list4) != 0:
                if lists[-1][0] != list4[0]:
                    a = lists.append(list4)
                    d1 = datetime.datetime.strptime(lists[-2][1], '%Y-%m-%d %H:%M:%S')
                    d2 = datetime.datetime.strptime(lists[-1][1], '%Y-%m-%d %H:%M:%S')
                    d = d2 - d1
                    h = int(format(d.seconds)) / 3600
                    # print(format(d.seconds))
                    list1 = lists[-2][0]
                    list2 = lists[-1][0]
                    list1 = list(map(float, list1))
                    list2 = list(map(float, list2))
                    list1[0], list1[1] = list1[1], list1[0]
                    list2[0], list2[1] = list2[1], list2[0]
                    tup1 = tuple(list1)
                    tup2 = tuple(list2)
                    # 两个经纬度之间的距离
                    km = geodesic(tup1, tup2).km
                    v = km / h
                    v_list.append(v)
                    if h>4:
                        fatigue_list.append(h)
                    if v>120:
                        vv_list.append(v)
                    time_list.append(d.seconds)
                    a = (v_list[-1] - v_list[-2]) / time_list[-1]
                    temp_a=abs(a)
                    if temp_a>20:
                        a_list.append(a)
                    if int(lists[-1][-1]) in Interval(10, 45) and v_list[-1]>100:
                        changLane_list.append(1)
                    elif int(lists[-1][-1]) in Interval(46, 90) and v_list[-1]>80:
                        changLane_list.append(2)
                    elif int(lists[-1][-1]) in Interval(91, 180) and v_list[-1]>60:
                        changLane_list.append(3)
                    elif int(lists[-1][-1]) in Interval(181, 270) and v_list[-1]>40:
                        changLane_list.append(4)
                    elif int(lists[-1][-1]) in Interval(271, 360) and v_list[-1]>30:
                        changLane_list.append(5)
        sizes=[len(FatigueDriving_time_list),len(vv_list),len(ldleSpeed_time_list),len(changLane_list),len(flameout_list),len(a_list)]
        name=['FatigueDriving','Speeding','ldleSpeed','changLane','flameout','Accelerated']
        explode = (0,0,0,0,0,0)
        plt.figure()
        plt.title(fileDir[17:24])
        plt.pie(sizes, labels=name, autopct="%.2f%%", shadow=False, explode=explode)
        plt.tight_layout()
        plt.savefig('F:\\pycharm\\jiqi\\' + '不良驾驶行为饼图'+fileDir[17:24] + '.png')
        plt.show()
        print(str(fileDir) + "不良驾驶行为统计数据:" +"--疲劳驾驶次数：" + str(len(FatigueDriving_time_list))+"--超速行驶次数："+str(len(vv_list))+"--急加速急减速次数："+str(len(a_list))+"--急变道次数："+str(len(changLane_list))+"--熄火滑行次数："+str(len(flameout_list))+"--怠速次数："+str(len(ldleSpeed_time_list)))
        wb = Workbook()
        ws = wb.active
        ws.title = fileDir[17:24]
        for i in range(len(name)):
            ws.cell(row=1, column=i+1).value = name[i]
        for i in range(len(FatigueDriving_time_list)):
            ws.cell(row=i+2, column=1).value = FatigueDriving_time_list[i]
        for i in range(len(vv_list)):
            ws.cell(row=i+2, column=2).value = vv_list[i]
        for i in range(len(ldleSpeed_time_list)):
            ws.cell(row=i+2, column=3).value = ldleSpeed_time_list[i]
        for i in range(len(changLane_list)):
            ws.cell(row=i+2, column=4).value = changLane_list[i]
        for i in range(len(flameout_list)):
            ws.cell(row=i+2, column=5).value = flameout_list[i]
        for i in range(len(a_list)):
            ws.cell(row=i+2, column=6).value = a_list[i]
        wb.save('F:\\pycharm\\jiqi\\' + '不良驾驶行为具体数据'+fileDir[17:24] + '.xlsx')
if __name__ == '__main__':
    ff = ['AA00001', 'AB00006', 'AD00003', 'AD00013', 'AD00053', 'AD00083', 'AD00419', 'AF00098', 'AF00131', 'AF00373']
    route_lists=[]
    for i in ff:
        fileDir = 'F:\\pycharm\\shuju\\' + i + '.csv'
        route(fileDir)

